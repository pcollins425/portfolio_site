"""Warehouse inventory read API — COMPINFO landing rows at warehouse properties."""

from __future__ import annotations

import math
import os
from datetime import date, datetime
from io import BytesIO

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app import mssql

router = APIRouter(prefix="/api/warehouse-inventory", tags=["warehouse-inventory"])

_WHERE_WAREHOUSE = """
    LOWER(LTRIM(RTRIM(ISNULL(property, N'')))) LIKE N'%warehouse%'
"""


def _catalog() -> str:
    return (os.environ.get("MSSQL_DATABASE") or "dgs_application_db").strip()


def _field_query(sql: str, params=None):
    return mssql.query(
        sql,
        params=params,
        database=_catalog(),
        profile="field",
        load_env=False,
    )


def _json_value(v):
    if isinstance(v, datetime):
        return v.date().isoformat() if v else None
    if isinstance(v, date):
        return v.isoformat()
    if v is None:
        return None
    return str(v).strip() if isinstance(v, str) else v


def _date_received(row: dict) -> str | None:
    for key in ("purch_date", "adddate"):
        val = _json_value(row.get(key))
        if val:
            return val
    return None


def _cabinet_label(manufacturer: str | None, cabinet: str | None) -> str:
    m = (manufacturer or "").strip()
    c = (cabinet or "").strip()
    if m and c:
        return f"{m} {c}"
    return m or c or "—"


def _fetch_warehouse_totals() -> list[dict]:
    rows = _field_query(
        f"""
        SELECT LTRIM(RTRIM(property)) AS property, COUNT(*) AS total
        FROM inventory.compinfo_landing
        WHERE {_WHERE_WAREHOUSE}
        GROUP BY LTRIM(RTRIM(property))
        ORDER BY COUNT(*) DESC, LTRIM(RTRIM(property))
        """
    )
    return [{"property": r["property"], "total": int(r["total"])} for r in rows]


def _sort_key(value: str | None) -> str:
    return (value or "").casefold()


def _build_pivot_data() -> dict:
    warehouses = _fetch_warehouse_totals()

    rows_raw = _field_query(
        f"""
        SELECT
            LTRIM(RTRIM(ISNULL(manufac, N''))) AS manufac,
            LTRIM(RTRIM(ISNULL(model_no, N''))) AS model_no,
            LTRIM(RTRIM(property)) AS property,
            COUNT(*) AS total
        FROM inventory.compinfo_landing
        WHERE {_WHERE_WAREHOUSE}
        GROUP BY
            LTRIM(RTRIM(ISNULL(manufac, N''))),
            LTRIM(RTRIM(ISNULL(model_no, N''))),
            LTRIM(RTRIM(property))
        """
    )

    pivot: dict[tuple[str, str], dict] = {}
    for r in rows_raw:
        manufac = r["manufac"] or ""
        model_no = r["model_no"] or ""
        prop = r["property"]
        count = int(r["total"])
        key = (manufac, model_no)
        if key not in pivot:
            pivot[key] = {
                "manufacturer": manufac or None,
                "cabinet": model_no or None,
                "label": _cabinet_label(manufac, model_no),
                "counts": {},
                "total": 0,
            }
        pivot[key]["counts"][prop] = count
        pivot[key]["total"] += count

    rows = sorted(
        pivot.values(),
        key=lambda x: (_sort_key(x["manufacturer"]), _sort_key(x["cabinet"])),
    )
    for row in rows:
        row["counts"] = {col["property"]: row["counts"].get(col["property"], 0) for col in warehouses}

    return {
        "grand_total": sum(w["total"] for w in warehouses),
        "columns": warehouses,
        "rows": rows,
    }


def _pivot_workbook(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Warehouse Inventory"

    columns = data["columns"]
    rows = data["rows"]
    grand_total = data["grand_total"]
    generated = datetime.now().strftime("%B %d, %Y")

    title_font = Font(bold=True, size=14)
    meta_font = Font(color="666666", size=11)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="5A85D6")
    group_font = Font(bold=True, color="5A85D6", size=11)
    group_fill = PatternFill("solid", fgColor="F4F7FB")
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill("solid", fgColor="FFF3EC")
    thin = Side(style="thin", color="D8DEE8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    indent = Alignment(horizontal="left", vertical="center", indent=1)

    last_col = 2 + len(columns) + 1
    last_letter = get_column_letter(last_col)

    ws["A1"] = "Warehouse Inventory — Cabinet by Warehouse"
    ws["A1"].font = title_font
    ws.merge_cells(f"A1:{last_letter}1")

    ws["A2"] = (
        f"Live from eMaint COMPINFO · {grand_total:,} assets across {len(columns)} warehouses"
        f" · Generated {generated}"
    )
    ws["A2"].font = meta_font
    ws.merge_cells(f"A2:{last_letter}2")

    header_row = 4
    headers = ["Manufacturer", "Cabinet"] + [col["property"] for col in columns] + ["Total"]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center if col_idx > 2 else left
        cell.border = border

    data_row = header_row + 1
    last_manufacturer = None
    col_totals = {col["property"]: 0 for col in columns}

    for row in rows:
        manufacturer = (row["manufacturer"] or "").strip() or "—"
        cabinet = (row["cabinet"] or "").strip() or "—"

        if manufacturer != last_manufacturer:
            group_cell = ws.cell(row=data_row, column=1, value=manufacturer)
            group_cell.font = group_font
            group_cell.fill = group_fill
            group_cell.alignment = left
            for col_idx in range(2, last_col + 1):
                cell = ws.cell(row=data_row, column=col_idx)
                cell.fill = group_fill
                cell.border = border
            ws.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=last_col)
            data_row += 1
            last_manufacturer = manufacturer

        ws.cell(row=data_row, column=1, value="").border = border
        cab_cell = ws.cell(row=data_row, column=2, value=cabinet)
        cab_cell.alignment = indent
        cab_cell.border = border

        for col_idx, col in enumerate(columns, start=3):
            count = row["counts"].get(col["property"], 0)
            col_totals[col["property"]] += count
            cell = ws.cell(row=data_row, column=col_idx, value=count or None)
            cell.alignment = center
            cell.border = border
            cell.number_format = "#,##0"

        total_cell = ws.cell(row=data_row, column=last_col, value=row["total"])
        total_cell.alignment = center
        total_cell.border = border
        total_cell.font = Font(bold=True)
        total_cell.number_format = "#,##0"
        data_row += 1

    for col_idx, col in enumerate(columns, start=3):
        cell = ws.cell(row=data_row, column=col_idx, value=col_totals[col["property"]])
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = center
        cell.border = border
        cell.number_format = "#,##0"

    total_label = ws.cell(row=data_row, column=2, value="All cabinets")
    total_label.font = total_font
    total_label.fill = total_fill
    total_label.alignment = left
    total_label.border = border
    ws.cell(row=data_row, column=1, value="").border = border

    grand_cell = ws.cell(row=data_row, column=last_col, value=grand_total)
    grand_cell.font = total_font
    grand_cell.fill = total_fill
    grand_cell.alignment = center
    grand_cell.border = border
    grand_cell.number_format = "#,##0"

    widths = {"A": 22, "B": 28}
    for col_idx, col in enumerate(columns, start=3):
        widths[get_column_letter(col_idx)] = max(12, min(18, len(col["property"]) + 2))
    widths[last_letter] = 10
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width

    ws.freeze_panes = f"A{header_row + 1}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _search_clause(search: str) -> tuple[str, tuple]:
    like = f"%{search.strip()}%"
    sql = """
        AND (
            serial_no LIKE %s
            OR manufac LIKE %s
            OR model_no LIKE %s
            OR prev_loca LIKE %s
        )
    """
    return sql, (like, like, like, like)


@router.get("/health")
def warehouse_inventory_health():
    catalog = _catalog()
    ok = False
    n = None
    try:
        row = _field_query(
            f"""
            SELECT COUNT(*) AS n
            FROM inventory.compinfo_landing
            WHERE {_WHERE_WAREHOUSE}
            """
        )[0]
        n = int(row["n"])
        ok = True
    except Exception:
        pass
    ext = os.environ.get("MSSQL_EXTERNAL") or os.environ.get("MSSQL_HOST")
    return {
        "ok": ok,
        "database": catalog,
        "warehouse_rows": n,
        "host": ext,
    }


@router.get("/summary")
def warehouse_summary():
    """Total assets per warehouse property (compinfo_landing)."""
    try:
        warehouses = _fetch_warehouse_totals()
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"database error: {exc}") from exc

    return {
        "warehouses": warehouses,
        "grand_total": sum(w["total"] for w in warehouses),
    }


@router.get("/pivot")
def warehouse_pivot():
    """CEO pivot: rows = manufacturer & cabinet, columns = all warehouses + total."""
    try:
        return _build_pivot_data()
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"database error: {exc}") from exc


@router.get("/export/pivot")
def warehouse_export_pivot():
    """Download the warehouse pivot as a formatted Excel workbook."""
    try:
        data = _build_pivot_data()
        content = _pivot_workbook(data)
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"database error: {exc}") from exc

    stamp = datetime.now().strftime("%Y-%m-%d")
    filename = f"warehouse-inventory-pivot-{stamp}.xlsx"
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/serials")
def warehouse_serials(
    property: str | None = Query(None, max_length=120, description="Warehouse property name"),
    manufacturer: str | None = Query(None, max_length=120),
    cabinet: str | None = Query(None, max_length=120),
    q: str = Query("", max_length=120, description="Search serial or previous location"),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=500),
):
    """Paginated serial list for pivot drill-down (cell, row, or column)."""
    search = q.strip()
    prop = property.strip() if property else None
    manufac = manufacturer.strip() if manufacturer else None
    model_no = cabinet.strip() if cabinet else None

    if not prop and not manufac and not model_no:
        raise HTTPException(status_code=400, detail="at least one filter is required")

    filter_sql = f"WHERE {_WHERE_WAREHOUSE}"
    params: list = []

    if prop:
        filter_sql += " AND LTRIM(RTRIM(property)) = %s"
        params.append(prop)

    if manufac is not None:
        filter_sql += " AND LTRIM(RTRIM(ISNULL(manufac, N''))) = %s"
        params.append(manufac)
    if model_no is not None:
        filter_sql += " AND LTRIM(RTRIM(ISNULL(model_no, N''))) = %s"
        params.append(model_no)

    search_sql = ""
    if search:
        search_sql, search_params = _search_clause(search)
        params.extend(search_params)

    try:
        count_row = _field_query(
            f"""
            SELECT COUNT(*) AS n
            FROM inventory.compinfo_landing
            {filter_sql}
            {search_sql}
            """,
            tuple(params),
        )[0]
        total_items = int(count_row["n"])

        offset = (page - 1) * page_size
        item_rows = _field_query(
            f"""
            SELECT
                serial_no,
                asset_id,
                manufac,
                model_no,
                property,
                prev_loca,
                purch_date,
                adddate
            FROM inventory.compinfo_landing
            {filter_sql}
            {search_sql}
            ORDER BY model_no, manufac, serial_no
            OFFSET {offset} ROWS FETCH NEXT {page_size} ROWS ONLY
            """,
            tuple(params),
        )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"database error: {exc}") from exc

    items = [
        {
            "serial": _json_value(r.get("serial_no")),
            "asset_id": _json_value(r.get("asset_id")),
            "manufacturer": _json_value(r.get("manufac")),
            "cabinet": _json_value(r.get("model_no")),
            "property": _json_value(r.get("property")),
            "date_received": _date_received(r),
            "previous_location": _json_value(r.get("prev_loca")) or None,
        }
        for r in item_rows
    ]

    total_pages = max(1, math.ceil(total_items / page_size)) if total_items else 1

    return {
        "property": prop,
        "manufacturer": manufac,
        "cabinet": model_no,
        "search": search or None,
        "total_items": total_items,
        "items": items,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
    }


@router.get("/warehouse")
def warehouse_detail(
    property: str = Query(..., min_length=1, max_length=120, description="Warehouse property name"),
    q: str = Query("", max_length=120, description="Search serial, manufacturer, cabinet, prev location"),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=500),
):
    """Cabinet breakdown + paginated asset list for one warehouse."""
    prop = property.strip()
    if not prop:
        raise HTTPException(status_code=400, detail="property is required")

    search = q.strip()

    filter_sql = f"""
        LTRIM(RTRIM(property)) = %s
        AND {_WHERE_WAREHOUSE}
    """
    search_sql = ""
    search_params: tuple = ()
    if search:
        search_sql, search_params = _search_clause(search)

    try:
        exists = _field_query(
            f"""
            SELECT COUNT(*) AS n
            FROM inventory.compinfo_landing
            WHERE {filter_sql}
            """,
            (prop,),
        )[0]
        if int(exists["n"]) == 0:
            raise HTTPException(status_code=404, detail=f"warehouse not found: {prop!r}")

        cabinet_rows = _field_query(
            f"""
            SELECT
                LTRIM(RTRIM(ISNULL(manufac, N''))) AS manufac,
                LTRIM(RTRIM(ISNULL(model_no, N''))) AS model_no,
                COUNT(*) AS total
            FROM inventory.compinfo_landing
            WHERE {filter_sql}
            {search_sql}
            GROUP BY LTRIM(RTRIM(ISNULL(manufac, N''))), LTRIM(RTRIM(ISNULL(model_no, N'')))
            ORDER BY COUNT(*) DESC, model_no, manufac
            """,
            (prop,) + search_params,
        )

        count_row = _field_query(
            f"""
            SELECT COUNT(*) AS n
            FROM inventory.compinfo_landing
            WHERE {filter_sql}
            {search_sql}
            """,
            (prop,) + search_params,
        )[0]
        total_items = int(count_row["n"])

        offset = (page - 1) * page_size
        item_rows = _field_query(
            f"""
            SELECT
                serial_no,
                manufac,
                model_no,
                prev_loca,
                purch_date,
                adddate
            FROM inventory.compinfo_landing
            WHERE {filter_sql}
            {search_sql}
            ORDER BY model_no, manufac, serial_no
            OFFSET {offset} ROWS FETCH NEXT {page_size} ROWS ONLY
            """,
            (prop,) + search_params,
        )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"database error: {exc}") from exc

    cabinet_counts = [
        {
            "manufacturer": r["manufac"] or None,
            "cabinet": r["model_no"] or None,
            "count": int(r["total"]),
        }
        for r in cabinet_rows
    ]

    items = [
        {
            "serial": _json_value(r.get("serial_no")),
            "manufacturer": _json_value(r.get("manufac")),
            "cabinet": _json_value(r.get("model_no")),
            "date_received": _date_received(r),
            "previous_location": _json_value(r.get("prev_loca")) or None,
        }
        for r in item_rows
    ]

    total_pages = max(1, math.ceil(total_items / page_size)) if total_items else 1

    return {
        "property": prop,
        "search": search or None,
        "total_assets": total_items,
        "distinct_cabinets": len(cabinet_counts),
        "cabinet_counts": cabinet_counts,
        "items": items,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
    }
